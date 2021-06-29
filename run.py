import os
from subprocess import call

# from numpy import flatiter
from db_operator import DBOperator, get_now_date
import pandas as pd
import openpyxl as px
from openpyxl.styles import PatternFill
from scrapy.crawler import CrawlerProcess
from scrapy.utils.project import get_project_settings


######################################################
# 共通定数
######################################################
PROPERTY_TYPE_INDEX = 0
PROPERTY_EVENT_INDEX = 1
PREFECTURE_INDEX = 2
LOCATION_NAME_INDEX = 3

######################################################
# 関数定義
######################################################


def adjust_cell_width(excel):
    if os.path.exists(excel):
        wb = px.load_workbook(filename=excel)
        sheet_num = len(wb.worksheets)
        for sheet_index in range(sheet_num):
            ws = wb.worksheets[sheet_index]
            set_column_width(ws)

        wb.save(excel)


def set_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def get_new_excel():
    excel = []

    for file in os.listdir():
        _, ext = os.path.splitext(file)
        if ext == '.xlsx':
            excel.append(file)
    if not excel:
        return None
    else:
        return max(excel)


def write_diff(before_file, after_file, filename):
    with pd.ExcelWriter(filename) as writer:
        # before,afterの全シートを読み込む
        df_sheet_all_before = pd.read_excel(before_file, sheet_name=None)
        df_sheet_all_after = pd.read_excel(after_file, sheet_name=None)

        for sheet in df_sheet_all_after.keys():
            is_sheet = sheet in df_sheet_all_before.keys()

            df_after = df_sheet_all_after[sheet]
            df_after = df_after.drop('Unnamed: 0', axis=1)
            df_after = df_after.drop('坪単価数値', axis=1)

            # 前回シートが存在した場合は物件番号が存在したかを判定する
            if is_sheet:
                df_before = df_sheet_all_before[sheet]
                df_after['diff'] = df_after['物件番号'].isin(df_before['物件番号'])
            df_after.to_excel(writer, sheet_name=sheet, index=False)


def fill_diff(excel):
    if os.path.exists(excel):
        wb = px.load_workbook(filename=excel)
        sheet_num = len(wb.worksheets)
        for sheet_index in range(sheet_num):
            ws = wb.worksheets[sheet_index]
            change_background_color(ws)

        wb.save(excel)


def search_diff_col(ws):
    is_diff = False
    for row in ws.iter_rows(max_row=1):
        col_index = 0
        for col in row:
            print(col.value)
            if col.value == 'diff':
                is_diff = True

                break
            else:
                col_index += 1

                continue
    return is_diff, col_index


def change_background_color(ws):
    fill = PatternFill(patternType='solid', fgColor='d3d3d3', bgColor='d3d3d3')
    is_diff, col_index = search_diff_col(ws)
    if not is_diff:
        return
    diff = []

    # diff列の結果を取得
    for cell in list(ws.columns)[col_index]:
        diff.append(cell.value)
    # 最初の行はカラム名のため必要なし
    diff.pop(0)

    row_list = range(2, len(diff) + 1)
    fill_list = []
    # 塗りつぶす行リストを作成
    for index, row in enumerate(row_list):
        if diff[index]:
            fill_list.append(row)

    for index in fill_list:
        # row = ws.row_dimensions[index]
        # row.fill = fill
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == index:
                    cell.fill = fill


def save_search_result(df, filename_tmp):
    mean = df['坪単価数値'].mean()
    df_want = df.query('坪単価数値 < @mean')
    df_want = df_want.sort_values('坪単価数値')
    if len(df_want) > 0:
        if os.path.exists(filename_tmp):
            with pd.ExcelWriter(filename_tmp, mode='a') as writer:
                df_want.to_excel(
                    writer,
                    sheet_name=area_key +
                    '_' +
                    property_event_key)
        else:
            with pd.ExcelWriter(filename_tmp) as writer:
                df_want.to_excel(
                    writer,
                    sheet_name=area_key +
                    '_' +
                    property_event_key)


######################################################
# メイン処理
######################################################
# print(os.getcwd())
# call(["scrapy", "crawl", "reins"])
# 物件情報抽出処理
process = CrawlerProcess(get_project_settings())

process.crawl('reins', domain='system.reins.jp')
process.start()  # the script will block here until the crawling is finished

# 前処理
last_created_excel = get_new_excel()
df_list = pd.read_csv('property_search.csv')
filename = get_now_date() + '.xlsx'
filename_tmp = get_now_date() + '_tmp.xlsx'
# DBを読み込む
db = DBOperator('reins.db')


for _, query in df_list.iterrows():
    property_event_key = query[PROPERTY_EVENT_INDEX]
    area_key = query[LOCATION_NAME_INDEX]

    # 所在地が空の場合NAN(float)が格納されている
    if isinstance(property_event_key, float):
        property_event_key = query[PROPERTY_TYPE_INDEX]

    if isinstance(area_key, float):
        area_key = query[PREFECTURE_INDEX]

    df = db.find_post(area_key, property_event_key)
    save_search_result(df, filename_tmp)

if last_created_excel is not None:
    write_diff(last_created_excel, filename_tmp, filename)
    fill_diff(filename)

    # tmpのexcelファイルを削除
    os.remove(filename_tmp)

    # # セル幅調整
    adjust_cell_width(filename)
